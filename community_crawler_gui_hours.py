import os, re, sys, time, threading, random, json, base64, platform, uuid
from datetime import datetime, timedelta
from urllib.parse import urlparse, urljoin, urlunparse, urlencode, parse_qs
import pandas as pd

# GUI
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Excel 워터마킹
from openpyxl import load_workbook

# Selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# RSA 공개키 검증
try:
    from cryptography.hazmat.primitives import serialization, hashes
    from cryptography.hazmat.primitives.asymmetric import padding
except Exception:
    raise RuntimeError("필수 모듈 'cryptography'가 없습니다. 설치:  pip install cryptography")

APP_TITLE = "커뮤니티 크롤러 (오프라인 라이선스 + 워터마킹)"
USER_HOME = os.path.expanduser("~")
DEFAULT_DESKTOP = os.path.join(USER_HOME, "Desktop")
APP_DIR = os.path.join(os.getenv("APPDATA") or USER_HOME, "CommunityCrawler")
LICENSE_PATH = os.path.join(APP_DIR, "license.lic")

# 내부 안전 한도
MAX_PAGES_SOFT   = 50
STALE_PAGE_LIMIT = 3

# ====== [중요] 공개키를 여기에 붙여주세요 ======
PUBLIC_PEM = b"""-----BEGIN PUBLIC KEY-----
MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEArwh9qGLUP3alVE/keAHz
dV53lBIEVpzzuvTpi/EPXqufIXdjfGupZbpF8M7yUGtdsD8WGpW27BKuR4FQQmPO
SNp6lIPwKlTvn46Y3R/nHFE9s0WazUyWIa7mkA0DbMhTihP6x7Lq2Y0dmEZUTJm0
mKEzG+YF6RwOEmctHG05YqyK7xZEzSNNXK2m3hSCptf4romsrty5Hh64vsZ1nR4Z
rNc3zdmMO4MZFWlccDQRpgvDmTj/+IqbQnsfMdPy8FoW8Wm/zPhKQQ22J1LXirnX
5NoWhclvGNy2i4llOP26cNrvK+s5juGKJGhWe698LnrQZLMtzT27px/oqS7n14Ya
zwIDAQAB
-----END PUBLIC KEY-----
"""
# ==============================================

def ts():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def ensure_dir_for_file(path: str):
    d = os.path.dirname(os.path.abspath(path))
    if d: os.makedirs(d, exist_ok=True)

def default_xlsx_path():
    return os.path.join(DEFAULT_DESKTOP, f"크롤링_결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

def to_int_or_none(text):
    try: return int(re.sub(r"[^\d]", "", str(text)))
    except Exception: return None

def add_or_replace_query_param(url: str, key: str, value) -> str:
    parts = list(urlparse(url))
    q = parse_qs(parts[4], keep_blank_values=True)
    q[key] = [str(value)]
    parts[4] = urlencode(q, doseq=True)
    return urlunparse(parts)

def rsleep(min_s=0.1, max_s=0.5):
    time.sleep(random.uniform(min_s, max_s))

# ---------------- 라이선스(오프라인, 공개키 서명) ----------------
def machine_id():
    try:
        if platform.system() == "Windows":
            import winreg
            k = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Cryptography")
            v, _ = winreg.QueryValueEx(k, "MachineGuid")
            return v
        elif platform.system() == "Darwin":
            import subprocess
            out = subprocess.check_output(["ioreg","-rd1","-c","IOPlatformExpertDevice"]).decode(errors="ignore")
            return out.split('IOPlatformUUID" = "')[1].split('"')[0]
        else:
            return open("/etc/machine-id").read().strip()
    except Exception:
        return str(uuid.getnode())

def b64u_decode(s: str) -> bytes:
    # urlsafe padding 보정
    pad = "=" * (-len(s) % 4)
    return base64.urlsafe_b64decode(s + pad)

def verify_license_text(lic_text: str):
    """ lic_text(json): {"payload": b64(json), "sig": b64(signature)} """
    try:
        lic = json.loads(lic_text)
        payload_json = b64u_decode(lic["payload"])
        payload = json.loads(payload_json)
        sig = b64u_decode(lic["sig"])

        pub = serialization.load_pem_public_key(PUBLIC_PEM)
        pub.verify(sig, payload_json, padding.PKCS1v15(), hashes.SHA256())

        # 기기락 + 만료 체크 (로컬 시간 기준)
        dev = payload.get("dev") or ""
        exp = payload.get("exp") or ""
        user = payload.get("user") or ""

        if dev and dev != machine_id():
            return False, "등록된 PC가 아닙니다.", None
        if exp:
            if datetime.now().date() > datetime.strptime(exp, "%Y-%m-%d").date():
                return False, "라이선스가 만료되었습니다.", None

        return True, "", payload  # payload: {"user","dev","exp",...}
    except Exception as e:
        return False, f"라이선스 검증 실패: {e}", None

def load_license_from_disk():
    # 1) AppData 경로
    candidates = [LICENSE_PATH]
    # 2) 실행 파일과 동일한 경로 (포터블)
    base = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
    candidates.append(os.path.join(base, "license.lic"))
    for path in candidates:
        try:
            return open(path, "r", encoding="utf-8").read()
        except Exception:
            continue
    return None

def save_license_to_disk(text: str):
    os.makedirs(APP_DIR, exist_ok=True)
    with open(LICENSE_PATH, "w", encoding="utf-8") as f:
        f.write(text)

def select_and_verify_license(parent) -> dict | None:
    path = filedialog.askopenfilename(parent=parent, title="라이선스 파일(.lic) 선택",
                                      filetypes=[("License file","*.lic"),("All files","*.*")])
    if not path: return None
    txt = open(path, "r", encoding="utf-8").read()
    ok, msg, payload = verify_license_text(txt)
    if not ok:
        messagebox.showerror("라이선스 오류", msg)
        return None
    save_license_to_disk(txt)
    return payload

def ensure_license(parent) -> dict | None:
    # 1) 디스크에 있으면 검증
    txt = load_license_from_disk()
    if txt:
        ok, msg, payload = verify_license_text(txt)
        if ok: return payload
        messagebox.showwarning("라이선스 재인증", f"{msg}\n새 라이선스를 선택하세요.")
    # 2) 파일 선택 유도
    return select_and_verify_license(parent)

# ---------------- 워터마킹(엑셀 숨김 시트) ----------------
def watermark_excel(path: str, payload: dict | None):
    if not payload: return
    try:
        wb = load_workbook(path)
        ws = wb.create_sheet("_meta")
        ws.sheet_state = "hidden"
        ws["A1"], ws["B1"] = "user",   payload.get("user","")
        ws["A2"], ws["B2"] = "device", payload.get("dev","")
        ws["A3"], ws["B3"] = "exp",    payload.get("exp","")
        wb.save(path)
    except Exception as e:
        # 워터마크 실패는 치명적이지 않으므로 로그만
        print("워터마크 실패:", e)

# ---------------- Selenium 공통 ----------------
def initialize_driver(show_browser: bool):
    os.environ.setdefault("WDM_LOG_LEVEL", "0")
    options = Options()
    if not show_browser:
        options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    try:
        driver_path = ChromeDriverManager().install()
        service = Service(driver_path)
    except Exception as e:
        base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
        local_driver = os.path.join(base_dir, "chromedriver.exe")
        if not os.path.exists(local_driver):
            raise RuntimeError(
                "ChromeDriver 자동 설치 실패: {}\n"
                "→ 네트워크/방화벽 확인 또는 chromedriver.exe를 실행 폴더에 두세요."
                .format(e)
            )
        service = Service(local_driver)
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(25)
    return driver

# ---------------- 날짜 파싱 유틸 ----------------
_DOT_DT_RE = re.compile(r"^(\d{4})\.(\d{2})\.(\d{2})\s+(\d{2}):(\d{2})$")
def parse_dt_dot(text: str):
    if not text: return None
    m = _DOT_DT_RE.match(text.strip())
    if not m: return None
    y,M,d,h,mi = map(int, m.groups())
    try: return datetime(y,M,d,h,mi)
    except ValueError: return None

_HHMM_RE = re.compile(r"^(\d{1,2}):(\d{2})$")
def parse_dt_hhmm_today(text: str):
    if not text: return None
    m = _HHMM_RE.match(text.strip())
    if not m: return None
    h,mi = map(int, m.groups())
    now = datetime.now()
    try: return datetime(now.year, now.month, now.day, h, mi)
    except ValueError: return None

def parse_dt_dc_flexible(text: str):
    if not text: return None
    s = text.strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})\s+(\d{2}):(\d{2})(?::(\d{2}))?$", s)
    if m:
        y,M,d,h,mi,sec = m.groups()
        try: return datetime(int(y),int(M),int(d),int(h),int(mi),int(sec or "0"))
        except ValueError: return None
    return parse_dt_hhmm_today(s)

# ---------------- FMKorea ----------------
FM_LINK_PATTERNS = [
    re.compile(r"/\d{5,}$"),
    re.compile(r"[?&]document_srl=\d+"),
]

def fmk_collect_links_by_user_selector(driver):
    sel = '.pc_voted_count.pc_voted_count_plus.pc_voted_count_short'
    cand = driver.find_elements(By.CSS_SELECTOR, sel)
    links, seen = [], set()
    for el in cand:
        href = el.get_attribute('href')
        if href and href not in seen:
            seen.add(href); links.append(href)
    return links

def collect_links_fallback_regex(driver):
    base = driver.current_url
    links, seen = [], set()
    for a in driver.find_elements(By.CSS_SELECTOR, "a[href]"):
        href = a.get_attribute("href")
        if not href: continue
        abs_href = urljoin(base, href)
        if any(p.search(abs_href) for p in FM_LINK_PATTERNS):
            if abs_href not in seen:
                seen.add(abs_href); links.append(abs_href)
    return links

def fmk_get_content(link, driver):
    driver.get(link); rsleep()
    try:
        wait = WebDriverWait(driver, 5)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".np_18px_span")))
        poten = driver.find_elements(By.CSS_SELECTOR, "h1.np_18px > span.STAR-BEST_T")
        title_el = driver.find_elements(By.CSS_SELECTOR, ".np_18px_span")
        title = (title_el[0].text.strip() if title_el else "제목 없음")
        if poten and title_el: title = f"포텐: {title}"
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".date.m_no")))
        date_text = driver.find_element(By.CSS_SELECTOR, ".date.m_no").text.strip()
        wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '조회 수')]/b")))
        views_text = driver.find_element(By.XPATH, "//span[contains(text(), '조회 수')]/b").text.strip()
        views = to_int_or_none(views_text)
    except Exception as e:
        print("FMK 상세 파싱 오류:", e)
        title, date_text, views = "제목 없음", "", None
    return title, date_text, views

def crawl_fmkorea(list_url, cutoff, show_browser, log):
    rows = []
    driver = initialize_driver(show_browser)
    try:
        page, stale_pages = 1, 0
        while page <= MAX_PAGES_SOFT:
            url = add_or_replace_query_param(list_url, "page", page)
            log(f"[FMK] 목록 page={page} | {url}")
            driver.get(url); rsleep()

            links = fmk_collect_links_by_user_selector(driver) or collect_links_fallback_regex(driver)
            log(f"[FMK] 후보 {len(links)}개")
            if not links:
                stale_pages += 1
                if stale_pages >= STALE_PAGE_LIMIT: log("[FMK] 연속 없음 → 종료"); break
                page += 1; continue
            stale_pages = 0

            found_old = False
            for href in links:
                title, date_text, views = fmk_get_content(href, driver); rsleep()
                dt = parse_dt_dot(date_text)
                if not dt:
                    log(f"[FMK] 날짜 파싱 실패 → 스킵: {date_text} | {href}")
                    continue
                rows.append({
                    "Site":"FMKorea","Title":title,"Date":date_text,
                    "DateISO": dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "Views":views, "Link":href
                })
                if dt < cutoff: found_old = True
            if found_old:
                log("[FMK] 오래된 글 감지 → 이 페이지까지 수집 후 종료"); break
            page += 1
    finally:
        driver.quit()
    return rows

# ---------------- DCInside ----------------
def crawl_dcinside(list_url, cutoff, show_browser, log):
    rows = []
    driver = initialize_driver(show_browser)
    log(f"[DC] cutoff = {cutoff:%Y-%m-%d %H:%M:%S}")
    try:
        stale_pages = 0
        for page in range(1, MAX_PAGES_SOFT+1):
            url = add_or_replace_query_param(list_url, "page", page)
            log(f"[DC] 목록 page={page} | {url}")
            driver.get(url); rsleep()
            base = driver.current_url
            trs = driver.find_elements(By.CSS_SELECTOR, "tr.ub-content.us-post")
            log(f"[DC] 행 {len(trs)}")
            if not trs:
                stale_pages += 1
                if stale_pages >= STALE_PAGE_LIMIT: log("[DC] 연속 없음 → 종료"); break
                continue
            stale_pages = 0
            found_recent = False
            for tr in trs:
                try:
                    a = tr.find_element(By.CSS_SELECTOR, "td.gall_tit a[href]")
                    href = urljoin(base, a.get_attribute("href"))
                    title = a.text.strip() or (a.get_attribute("title") or "").strip()
                    d = tr.find_element(By.CSS_SELECTOR, "td.gall_date")
                    date_text = (d.get_attribute("title") or d.text or "").strip()
                    dt = parse_dt_dc_flexible(date_text)
                    v = tr.find_element(By.CSS_SELECTOR, "td.gall_count")
                    views = to_int_or_none(v.text)
                    if dt and dt >= cutoff:
                        rows.append({
                            "Site":"DCInside","Title":title or "제목 없음",
                            "Date":date_text,"DateISO":dt.strftime("%Y-%m-%d %H:%M:%S"),
                            "Views":views,"Link":href
                        }); found_recent = True
                except Exception as e:
                    log(f"[DC] 행 파싱 실패: {e}")
            if not found_recent:
                stale_pages += 1
                if stale_pages >= STALE_PAGE_LIMIT: log("[DC] 최근 글 없음 연속 → 종료"); break
            else:
                stale_pages = 0
    finally:
        driver.quit()
    return rows

# ---------------- TheQoo (상세 + 공지 제외 + .side.fr span + 조회수 count_container) ----------------
_DOT_FULL_RE = re.compile(r"^(\d{4})\.(\d{2})\.(\d{2})\s+(\d{2}):(\d{2})$")
_DOT_Y2_RE   = re.compile(r"^(\d{2})\.(\d{2})\.(\d{2})$")
_DOT_MD_RE   = re.compile(r"^(\d{2})\.(\d{2})$")
def parse_dt_theqoo(text: str):
    if not text: return None
    s = text.strip()
    m = _DOT_FULL_RE.match(s)
    if m:
        y,M,d,h,mi = map(int, m.groups())
        try: return datetime(y,M,d,h,mi)
        except ValueError: return None
    m = _DOT_Y2_RE.match(s)   # 24.12.06
    if m:
        yy,M,d = map(int, m.groups())
        y = 2000 + yy
        try: return datetime(y,M,d,0,0)
        except ValueError: return None
    m = _DOT_MD_RE.match(s)   # 08.15
    if m:
        M,d = map(int, m.groups()); now=datetime.now()
        try: return datetime(now.year,M,d,0,0)
        except ValueError: return None
    m = _HHMM_RE.match(s)     # 12:39
    if m:
        h,mi = map(int, m.groups()); now=datetime.now()
        try: return datetime(now.year,now.month,now.day,h,mi)
        except ValueError: return None
    return None

def theqoo_collect_detail_links(driver):
    base = driver.current_url
    links, seen = [], set()
    title_tds = driver.find_elements(By.CSS_SELECTOR, "td.title")
    for td in title_tds:
        try:
            tr = td.find_element(By.XPATH, "./ancestor::tr[1]")
            # 공지 제외
            try:
                no_strong = tr.find_element(By.CSS_SELECTOR, "td.no strong")
                if "공지" in (no_strong.text or "").strip():
                    continue
            except Exception:
                pass
            a = td.find_element(By.CSS_SELECTOR, "a[href]:not(.replyNum)")
            href = urljoin(base, a.get_attribute("href"))
            if href not in seen:
                seen.add(href); links.append(href)
        except Exception:
            continue
    return links

def theqoo_parse_detail(driver, url):
    driver.get(url); rsleep()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))

    # 제목
    title = ""
    for sel in ["h1.title", ".title h1", ".title", "h1", "h2"]:
        els = driver.find_elements(By.CSS_SELECTOR, sel)
        if els and els[0].text.strip():
            title = els[0].text.strip(); break
    if not title: title = "제목 없음"

    # 날짜(.side.fr span 우선)
    date_text = ""
    for sel in [".side.fr span", ".date", ".regdate", ".time", "time[datetime]"]:
        els = driver.find_elements(By.CSS_SELECTOR, sel)
        if els:
            t = (els[0].get_attribute("datetime") or els[0].text or "").strip()
            if t: date_text = t; break
    if not date_text:
        m = re.search(r"\d{4}\.\d{2}\.\d{2}\s+\d{2}:\d{2}", driver.page_source)
        if m: date_text = m.group(0)

    # 조회수: .count_container 텍스트에서 첫 숫자
    views = None
    try:
        cnt = driver.find_element(By.CSS_SELECTOR, ".count_container")
        raw = (cnt.get_attribute("innerText") or cnt.text or "").strip()
        nums = re.findall(r"\d{1,3}(?:,\d{3})*|\d+", raw)
        if nums:
            views = to_int_or_none(nums[0])
    except NoSuchElementException:
        pass
    if views is None:
        # 백업: 페이지 전체 숫자에서 최대값 추정(원치 않으면 제거)
        all_nums = re.findall(r"\d{1,3}(?:,\d{3})*|\d+", driver.page_source)
        if all_nums:
            views = max((to_int_or_none(n) for n in all_nums), default=None)

    # 날짜 파싱
    dt = parse_dt_dot(date_text) or parse_dt_theqoo(date_text)

    return {
        "Site": "TheQoo",
        "Title": title,
        "Date": date_text,
        "DateISO": dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "",
        "Views": views,
        "Link": url,
        "_dt": dt
    }

def crawl_theqoo(list_url, cutoff, show_browser, log):
    rows = []
    driver = initialize_driver(show_browser)
    try:
        page, stale_pages = 1, 0
        while page <= MAX_PAGES_SOFT:
            url = add_or_replace_query_param(list_url, "page", page)
            log(f"[TQ] 목록 page={page} | {url}")
            driver.get(url); rsleep()

            links = theqoo_collect_detail_links(driver)
            log(f"[TQ] 상세 후보(공지 제외) {len(links)}개")
            if not links:
                stale_pages += 1
                if stale_pages >= STALE_PAGE_LIMIT: log("[TQ] 연속 없음 → 종료"); break
                page += 1; continue
            stale_pages = 0

            found_old = False
            for i, href in enumerate(links, 1):
                try:
                    post = theqoo_parse_detail(driver, href); rsleep()
                    dt = post["_dt"]
                    rows.append({
                        "Site": post["Site"], "Title": post["Title"],
                        "Date": post["Date"], "DateISO": post["DateISO"],
                        "Views": post["Views"], "Link": post["Link"]
                    })
                    if dt and dt < cutoff: found_old = True
                    if i % 10 == 0 or i == len(links):
                        log(f"[TQ] 진행 {i}/{len(links)} (누적 {len(rows)})")
                except Exception as e:
                    log(f"[TQ] 상세 실패: {e}")

            if found_old:
                log("[TQ] 오래된 글 감지 → 이 페이지까지 수집 후 종료"); break
            page += 1
    finally:
        driver.quit()
    return rows

# ---------------- GUI ----------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("880x700")
        self.resizable(False, False)

        self.var_comm    = tk.StringVar(value="FMKorea")   # FMKorea / DCInside / TheQoo
        self.var_url     = tk.StringVar(value="")
        self.var_days    = tk.IntVar(value=1)
        self.var_hours   = tk.IntVar(value=0)
        self.var_out     = tk.StringVar(value=default_xlsx_path())
        self.var_show    = tk.BooleanVar(value=False)

        self.license_payload = None  # {"user","dev","exp",...}

        self._build_ui()
        # 시작 시 라이선스 확인(없으면 선택)
        self.after(200, self._check_license_on_start)

    def _build_ui(self):
        pad = {'padx': 8, 'pady': 6}
        root = ttk.Frame(self); root.pack(fill="both", expand=True)

        # 라이선스 상태 표시줄
        self.lbl_license = ttk.Label(root, text="라이선스: 확인 중...", foreground="#444")
        self.lbl_license.grid(row=0, column=0, columnspan=4, sticky="w", padx=8, pady=(8,0))

        ttk.Label(root, text="1) 커뮤니티 선택").grid(row=1, column=0, sticky="w", **pad)
        ttk.Combobox(root, textvariable=self.var_comm,
                     values=["FMKorea","DCInside","TheQoo"],
                     state="readonly", width=20).grid(row=2, column=0, sticky="w", **pad)

        ttk.Label(root, text="2) 목록 URL").grid(row=3, column=0, sticky="w", **pad)
        ttk.Entry(root, textvariable=self.var_url, width=106).grid(row=4, column=0, columnspan=4, sticky="w", **pad)

        row3 = ttk.Frame(root); row3.grid(row=5, column=0, columnspan=4, sticky="w", **pad)
        ttk.Label(row3, text="3) 최근 기간").grid(row=0, column=0, sticky="w", padx=(0,6))
        ttk.Label(row3, text="일").grid(row=0, column=2, sticky="w", padx=(6,6))
        ttk.Label(row3, text="시간").grid(row=0, column=4, sticky="w", padx=(6,6))
        ttk.Spinbox(row3, from_=0, to=365, textvariable=self.var_days, width=5).grid(row=0, column=1, sticky="w")
        ttk.Spinbox(row3, from_=0, to=23,  textvariable=self.var_hours, width=5).grid(row=0, column=3, sticky="w", padx=(0,12))
        ttk.Checkbutton(row3, text="크롤링 화면 보기(브라우저 표시)", variable=self.var_show).grid(row=0, column=5, sticky="w")

        ttk.Label(root, text="4) 엑셀 저장 경로").grid(row=6, column=0, sticky="w", **pad)
        row_out = ttk.Frame(root); row_out.grid(row=7, column=0, columnspan=4, sticky="w", **pad)
        ttk.Entry(row_out, textvariable=self.var_out, width=92).grid(row=0, column=0, sticky="w", padx=(0,6))
        ttk.Button(row_out, text="찾아보기…", command=self.pick_out_path).grid(row=0, column=1, sticky="w")

        btns = ttk.Frame(root); btns.grid(row=8, column=0, columnspan=4, sticky="w", **pad)
        ttk.Button(btns, text="라이선스 불러오기", command=self.on_license_load).grid(row=0, column=0, padx=(0,8))
        ttk.Button(btns, text="실행", command=self.on_run).grid(row=0, column=1, padx=(0,8))
        ttk.Button(btns, text="종료", command=self.destroy).grid(row=0, column=2)

        ttk.Label(root, text="로그").grid(row=9, column=0, sticky="w", **pad)
        self.txt = tk.Text(root, height=18, width=114)
        self.txt.grid(row=10, column=0, columnspan=4, sticky="w", padx=8, pady=(0,8))
        self.txt.configure(state="disabled")

        ttk.Label(root, text="원초적인사이트 데이터수집 프로그램").grid(row=11, column=0, columnspan=4, sticky="w", padx=8, pady=(0,8))

    def pick_out_path(self):
        path = filedialog.asksaveasfilename(
            title="엑셀 저장 경로", defaultextension=".xlsx",
            filetypes=[("Excel 파일","*.xlsx"),("모든 파일","*.*")],
            initialfile=os.path.basename(self.var_out.get() or default_xlsx_path()),
            initialdir=os.path.dirname(self.var_out.get() or default_xlsx_path()) or DEFAULT_DESKTOP
        )
        if path: self.var_out.set(path)

    def log(self, msg: str):
        self.txt.configure(state="normal")
        self.txt.insert("end", f"{ts()} | {msg}\n")
        self.txt.see("end"); self.txt.configure(state="disabled")
        self.update_idletasks()

    # ---- 라이선스 처리 ----
    def _check_license_on_start(self):
        payload = None
        txt = load_license_from_disk()
        if txt:
            ok, msg, payload = verify_license_text(txt)
            if not ok:
                self.lbl_license.configure(text=f"라이선스 오류: {msg}", foreground="#B71C1C")
            else:
                self.lbl_license.configure(text=f"라이선스 OK — {payload.get('user','')}", foreground="#1B5E20")
        else:
            self.lbl_license.configure(text="라이선스 없음 — [라이선스 불러오기]를 눌러 등록하세요.", foreground="#444")
        self.license_payload = payload

    def on_license_load(self):
        payload = select_and_verify_license(self)
        if payload:
            self.license_payload = payload
            self.lbl_license.configure(text=f"라이선스 OK — {payload.get('user','')}", foreground="#1B5E20")
            messagebox.showinfo("라이선스", "라이선스 등록 완료")

    def _require_license(self) -> bool:
        if self.license_payload: return True
        messagebox.showwarning("라이선스", "라이선스를 먼저 등록해주세요.")
        payload = select_and_verify_license(self)
        if payload:
            self.license_payload = payload
            self.lbl_license.configure(text=f"라이선스 OK — {payload.get('user','')}", foreground="#1B5E20")
            return True
        return False

    # ---- 실행 ----
    def on_run(self):
        if not self._require_license(): return

        comm  = self.var_comm.get().strip()
        url   = self.var_url.get().strip()
        days  = int(self.var_days.get()); hours = int(self.var_hours.get())
        outp  = self.var_out.get().strip()
        show  = bool(self.var_show.get())

        if not url:
            messagebox.showwarning("입력 확인","목록 URL을 입력하세요."); return
        if days < 0 or hours < 0 or hours > 23:
            messagebox.showwarning("입력 확인","일은 0 이상, 시간은 0~23 범위로 입력해주세요."); return
        total_hours = days*24 + hours
        if total_hours < 1:
            messagebox.showwarning("입력 확인","총 시간이 1시간 이상이어야 합니다."); return
        if not outp:
            outp = default_xlsx_path(); self.var_out.set(outp)

        host = urlparse(url).netloc.lower()
        if comm == "FMKorea" and "fmkorea.com" not in host:
            messagebox.showerror("오류","선택과 URL이 일치하지 않습니다(FMKorea)."); return
        if comm == "DCInside" and "dcinside.com" not in host:
            messagebox.showerror("오류","선택과 URL이 일치하지 않습니다(DCInside)."); return
        if comm == "TheQoo" and "theqoo.net" not in host:
            messagebox.showerror("오류","선택과 URL이 일치하지 않습니다(TheQoo)."); return

        cutoff = datetime.now() - timedelta(hours=total_hours)
        self.log(f"실행: {comm} | 최근 {days}일 {hours}시간 (총 {total_hours}시간) | 화면보기={show} | cutoff={cutoff:%Y-%m-%d %H:%M}")
        threading.Thread(target=self._crawl_and_save_safe,
                         args=(comm, url, cutoff, outp, show), daemon=True).start()

    def _crawl_and_save_safe(self, comm, url, cutoff, outp, show):
        try:
            if comm == "FMKorea":
                rows = crawl_fmkorea(url, cutoff, show, self.log)
            elif comm == "DCInside":
                rows = crawl_dcinside(url, cutoff, show, self.log)
            else:
                rows = crawl_theqoo(url, cutoff, show, self.log)

            if not rows:
                self.log("수집 결과가 비었습니다."); messagebox.showinfo("완료","수집 결과가 없습니다."); return

            # 엑셀 저장 (DateISO 제외)
            df = pd.DataFrame(rows)
            want = [c for c in ["Site","Title","Date","Views","Link"] if c in df.columns]
            df = df[want]
            ensure_dir_for_file(outp); df.to_excel(outp, index=False)

            # 워터마킹
            watermark_excel(outp, self.license_payload)

            # 수집된 시각 범위 로그
            try:
                dts = [datetime.strptime(r["DateISO"], "%Y-%m-%d %H:%M:%S")
                       for r in rows if r.get("DateISO")]
                if dts:
                    self.log(f"수집된 시각 범위: {min(dts):%Y-%m-%d %H:%M:%S} ~ {max(dts):%Y-%m-%d %H:%M:%S}")
            except Exception:
                pass

            self.log(f"완료! 저장: {outp} | 수집 {len(df)}건")
            messagebox.showinfo("완료", f"저장 완료\n{outp}\n총 {len(df)}건")
        except Exception as e:
            self.log(f"오류: {e}")
            messagebox.showerror("오류", str(e))


if __name__ == "__main__":
    app = App()
    app.mainloop()
